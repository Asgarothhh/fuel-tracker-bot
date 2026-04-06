import logging
from datetime import datetime, timezone, timedelta
from io import BytesIO
from src.app.bot.keyboards import BTN_ADMIN_DISPUTED, BTN_ADMIN_RECENT
from src.app.bot.notifications import send_operation_to_user
from aiogram import types, F
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from openpyxl import Workbook
from sqlalchemy import cast, String
from sqlalchemy.orm import joinedload  # Добавлен импорт для связи таблиц!

from src.app.belorusneft_api import fetch_operational_raw, parse_operations
from src.app.db import get_db_session
from src.app.models import FuelOperation, FuelCard, Car, User
from src.app.permissions import require_permission, user_has_permission
from src.app.bot.keyboards import (
    BTN_ADMIN_IMPORT,
    BTN_ADMIN_IMPORT_TEST,
    BTN_ADMIN_SCHEDULES,
    BTN_ADMIN_SCHEDULE_SET,
    BTN_ADMIN_SCHEDULE_DEL,
    BTN_ADMIN_USERS,
    BTN_ADMIN_PENDING,
    BTN_ADMIN_EXPORT_EXCEL,
    BTN_ADMIN_GEN_CODE,
    BTN_ADMIN_EXPORT_CODES,
    BTN_ADMIN_HELP,
)
from src.app.bot.utils import extract_args
from src.app.bot.handlers import admin_schedules as admin_schedules_mod
from src.app.bot.handlers import admin_users as admin_users_mod

logger = logging.getLogger(__name__)

ADMIN_HELP_TEXT = (
    "📖 *Справка администратора*\n\n"
    "*Импорт:*\n"
    "• «Обновить импорт» — загрузка операций за вчера из API.\n"
    "• «Тестовый импорт» — dry-run, смотрите логи.\n\n"
    "*Расписание:*\n"
    "/schedule_set имя HH:MM (UTC)\n"
    "/schedule_remove имя\n\n"
    "*Пользователи:*\n"
    "/users — список, /generate_code id, /export_codes\n\n"
    "*Операции:*\n"
    "• «Неподтверждённые» — очередь.\n"
    "/assign_op [op_id] [user_id]\n\n"
    "• «Экспорт в Excel» — выгрузка всех операций в файл."
)


def _ocr_text(op: FuelOperation) -> str:
    """Колонка OCR: сырой текст — в `raw_text_debug` (как в excel_export)."""
    if not op.ocr_data:
        return ""
    if isinstance(op.ocr_data, dict):
        return (
            op.ocr_data.get("raw_text_debug")
            or op.ocr_data.get("raw_text")
            or op.ocr_data.get("text")
            or ""
        )
    return str(op.ocr_data)

@require_permission("admin:manage")
async def cmd_admin_help(message: types.Message):
    await message.reply(ADMIN_HELP_TEXT, parse_mode="HTML")


@require_permission("admin:manage")
async def cmd_pending_ops(message: types.Message):
    with get_db_session() as db:
        ops = (
            db.query(FuelOperation)
            .options(joinedload(FuelOperation.presumed_user))  # Подгружаем связь с БД
            .filter(FuelOperation.status.in_(
                ["loaded_from_api", "pending", "requires_manual", "disputed", "rejected_by_other"]))
            .order_by(FuelOperation.id.desc())
            .limit(50)  # Увеличили лимит до 50
            .all()
        )

        if not ops:
            await message.reply("Список необработанных операций пуст.")
            return

        lines = []
        for o in ops:
            api = o.api_data if isinstance(o.api_data, dict) else {}
            row = api.get("row") if isinstance(api.get("row"), dict) else {}

            dt = o.date_time.strftime('%d.%m %H:%M') if o.date_time else "—"
            card = api.get("cardNumber") or row.get("cardNumber") or "—"
            qty = api.get("productQuantity") or row.get("productQuantity") or "0"

            # Извлекаем водителя из сырых данных Белоруснефти
            driver_api = api.get("driverName") or row.get("driverName") or "—"
            # Извлекаем пользователя из нашей базы (если бот смог его сопоставить)
            db_user = o.presumed_user.full_name if o.presumed_user else "❌ Не привязан"
            # Машина из API
            car_api = o.car_from_api or "—"

            # Русифицируем статусы для вывода в Telegram
            status_ru = {
                "loaded_from_api": "📥 Новая",
                "pending": "⏳ Ждет ответа",
                "requires_manual": "🛠 Ручная привязка",
                "disputed": "⚠️ Спорная",
                "rejected_by_other": "❌ Отклонена"
            }.get(o.status, o.status)

            # Формируем расширенную карточку
            line = (
                f"<code>#{o.id}</code> [{status_ru}] {dt} | 💳<code>{card[-4:]}</code> | ⛽️{qty}л\n"
                f"   👤 API: {driver_api} | БД: {db_user}\n"
                f"   🚗 Авто: {car_api}"
            )
            lines.append(line)

        header = f"<b>📋 Необработанные операции (показано {len(ops)}):</b>\n\n"
        full_text = header + "\n\n".join(lines)

        # Безопасная отправка (если текста больше 4000 символов, бьем на 2 сообщения)
        if len(full_text) > 4000:
            for x in range(0, len(full_text), 4000):
                await message.answer(full_text[x:x + 4000], parse_mode="HTML")
        else:
            await message.answer(full_text, parse_mode="HTML")


# ... (начало файла с импортами остается как в моем предыдущем ответе) ...

@require_permission("admin:manage")
async def btn_export_excel(message: types.Message):
    await message.answer("⏳ Формирую полный отчет по всем операциям...")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # СОЗДАЕМ ЛИСТЫ
    ws_cards = wb.create_sheet("Заправки_по_картам")
    ws_personal = wb.create_sheet("Заправки_личные_средства")
    ws_disputed = wb.create_sheet("Спорные заправки")

    headers = [
        "ID", "Тип", "Источник", "Дата", "Время", "Карта", "Топливо", "Литры", "Сумма", "АЗС", "Чек",
        "Авто(API)", "Водитель(API)", "OCR", "Предполагаемый", "Фактический", "Факт.Авто",
        "Инициатор", "Подтвердил", "Статус", "Дата подтв", "Прим", "Готовность"
    ]

    for ws in (ws_cards, ws_personal, ws_disputed):
        ws.append(headers)

    status_ru = {
        "loaded_from_api": "Загружена из API",
        "pending": "Ожидает подтверждения",
        "confirmed": "Подтверждена",
        "requires_manual": "Требует ручной обработки",
        "disputed": "Спорная",
        "rejected_by_other": "Отклонена",
        "import_error": "Ошибка импорта"
    }

    with get_db_session() as db:
        # ИЗМЕНЕНИЕ 1: Убираем .filter(...) по статусам, чтобы брать ВСЕ записи
        operations = db.query(FuelOperation).order_by(FuelOperation.date_time.desc()).all()

        if not operations:
            await message.answer("В базе данных нет операций для выгрузки.")
            return

        for op in operations:
            api = op.api_data if isinstance(op.api_data, dict) else {}
            row_inner = api.get("row") if isinstance(api.get("row"), dict) else {}
            card_o = api.get("card") if isinstance(api.get("card"), dict) else {}
            ocr = op.ocr_data if isinstance(op.ocr_data, dict) else {}

            if op.source == "personal_receipt":
                card = "—"
                pname = ocr.get("fuel_type") or "—"
                pq = ocr.get("quantity") if ocr.get("quantity") not in (None, "") else "—"
                cost = ocr.get("total_sum") if ocr.get("total_sum") not in (None, "") else "—"
                azs = ocr.get("azs_number") or "—"
                car_api = op.actual_car or "—"
                drv = "—"
            else:
                card = api.get("cardNumber") or card_o.get("cardNumber") or "—"
                pname = api.get("productName") or row_inner.get("productName") or "—"
                pq = api.get("productQuantity") or row_inner.get("productQuantity") or 0
                cost = api.get("productCost") or row_inner.get("productCost") or 0
                azs = api.get("azsNumber") or row_inner.get("azsNumber") or row_inner.get("AzsCode") or "—"
                car_api = op.car_from_api or api.get("carNum") or row_inner.get("carNum") or "—"
                drv = api.get("driverName") or row_inner.get("driverName") or "—"

            presumed = db.query(User).filter_by(id=op.presumed_user_id).first() if op.presumed_user_id else None
            confirmed = db.query(User).filter_by(id=op.confirmed_user_id).first() if op.confirmed_user_id else None

            fuel_type = "Топливная карта" if op.source == "api" else "Личные средства"
            dt = op.date_time

            is_personal = op.source == "personal_receipt"
            if is_personal:
                presumed_name = "—"
                confirmed_name = presumed.full_name if presumed else "—"
                final_confirmer = "—"
                confirmed_dt = "—"
            else:
                presumed_name = presumed.full_name if presumed else "—"
                confirmed_name = confirmed.full_name if confirmed else "—"
                final_confirmer = confirmed.full_name if confirmed else "—"
                confirmed_dt = op.confirmed_at.strftime("%d.%m.%Y %H:%M") if op.confirmed_at else "—"
            current_status = status_ru.get(op.status, op.status or "—")

            row_data = [
                op.id, fuel_type, op.source or "api",
                dt.strftime("%d.%m.%Y") if dt else "—",
                dt.strftime("%H:%M:%S") if dt else "—",
                card, pname, pq, cost, azs, op.doc_number or "—",
                car_api, drv, _ocr_text(op),
                presumed_name, confirmed_name, op.actual_car or car_api,
                presumed_name, final_confirmer,
                current_status,
                confirmed_dt,
                "", "Да" if op.status == "confirmed" else "Нет"
            ]

            # ИЗМЕНЕНИЕ 2: Новая логика распределения
            # Если статус проблемный — в "Спорные"
            if op.status in ("requires_manual", "rejected_by_other", "disputed", "import_error"):
                ws_disputed.append(row_data)
            # Все остальные (confirmed, pending, loaded_from_api) распределяем по источнику
            else:
                if op.source == "api":
                    ws_cards.append(row_data)
                else:
                    ws_personal.append(row_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = f"Full_Fuel_Report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=name),
        caption=f"✅ Отчет готов. Выгружено {len(operations)} записей."
    )


@require_permission("admin:manage")
async def cmd_run_import_now_dry(message: types.Message):
    from src.app.jobs import run_import_job

    try:
        run_import_job("manual_dry_run", dry_run=True)
        await message.reply("Тестовый импорт (dry-run) выполнен. Смотрите логи; БД не менялась.")
    except Exception as e:
        logger.exception("dry-run import")
        await message.reply(f"Ошибка: {e}")


@require_permission("admin:manage")
async def cmd_run_import_now(message: types.Message):
    date = datetime.now() - timedelta(days=1)
    new_count = 0
    try:
        raw = fetch_operational_raw(date)
        status = raw.get("status")
        json_payload = raw.get("json")
        debug_files = raw.get("debug_files") or []

        if status != 200:
            await message.reply(f"HTTP {status}. Debug: {', '.join(debug_files)}")
            return

        if not json_payload:
            await message.reply(f"Нет JSON. Debug: {', '.join(debug_files)}")
            return

        ops = parse_operations(json_payload)
        if not ops:
            await message.reply("В ответе 0 операций.")
            return

        with get_db_session() as db:
            for op in ops:
                # 1. Безопасно достаем вложенные словари
                op_row = op.get("row") if isinstance(op.get("row"), dict) else {}
                op_card = op.get("card") if isinstance(op.get("card"), dict) else {}
                # 2. Ищем дату (теперь проверяем и в корне, и в row)
                dt_raw = op.get("date_time") or op.get("dateTimeIssue") or op_row.get("dateTimeIssue")
                dt_obj = None
                if dt_raw:
                    try:
                        dt_obj = datetime.fromisoformat(str(dt_raw).replace("Z", "+00:00"))
                    except Exception:
                        pass

                # 3. Ищем чек (теперь проверяем и в корне, и в row)
                doc_raw = op.get("doc_number") or op.get("docNumber") or op_row.get("docNumber")
                doc = str(doc_raw).strip() if doc_raw is not None else None
                # 4. Ищем остальные данные (с учетом вложенности)
                driver_name = op.get("driverName") or op.get("driver_name") or op_row.get("driverName")
                card_num = op.get("cardNumber") or op.get("card_number") or op_card.get("cardNumber")
                if not card_num and isinstance(op.get("card"), (str, int)):
                    card_num = op.get("card")
                if card_num: card_num = str(card_num).strip()

                car_plate = op.get("carNum") or op.get("car_num") or op.get("car") or op_row.get("carNum")
                car_plate_norm = str(car_plate).strip().upper() if car_plate else None

                # ПУНКТ 6: АВТО-ДОБАВЛЕНИЕ ПОЛЬЗОВАТЕЛЕЙ ИЗ API (ДО ПРОВЕРКИ НА ДУБЛИ!)
                presumed_user = None
                if driver_name:
                    drv_clean = str(driver_name).strip()
                    presumed_user = db.query(User).filter(User.full_name.ilike(drv_clean)).first()
                    if not presumed_user:
                        presumed_user = User(full_name=drv_clean, active=True)
                        db.add(presumed_user)
                        db.flush()

                if not presumed_user and card_num:
                    presumed_user = db.query(User).filter(cast(User.cards, String).like(f"%{card_num}%")).first()

                if card_num:
                    fuel_card = db.query(FuelCard).filter_by(card_number=card_num).first()
                    if not fuel_card:
                        fuel_card = FuelCard(card_number=card_num, active=True)
                        db.add(fuel_card)
                        db.flush()
                    if presumed_user:
                        fuel_card.user_id = presumed_user.id
                        u_cards = list(presumed_user.cards or [])
                        if card_num not in u_cards:
                            u_cards.append(card_num)
                            presumed_user.cards = u_cards

                if car_plate_norm:
                    car_obj = db.query(Car).filter_by(plate=car_plate_norm).first()
                    if not car_obj:
                        car_obj = Car(plate=car_plate_norm)
                        db.add(car_obj)
                        db.flush()
                    if presumed_user:
                        owners = list(car_obj.owners or [])
                        if presumed_user.id not in owners:
                            owners.append(presumed_user.id)
                            car_obj.owners = owners
                        u_cars = list(presumed_user.cars or [])
                        if car_plate_norm not in u_cars:
                            u_cars.append(car_plate_norm)
                            presumed_user.cars = u_cars

                # Проверка дубликатов самой операции (после того как справочники пополнены)
                filters = [FuelOperation.source == "api"]
                if doc: filters.append(FuelOperation.doc_number == doc)
                if dt_obj: filters.append(FuelOperation.date_time == dt_obj)
                if db.query(FuelOperation).filter(*filters).first():
                    continue

                new_op = FuelOperation(
                    source="api",
                    api_data=op.get("raw") or op,
                    imported_at=datetime.now(timezone.utc),
                    status="loaded_from_api",
                )
                if doc: new_op.doc_number = doc
                if dt_obj: new_op.date_time = dt_obj
                if presumed_user: new_op.presumed_user_id = presumed_user.id
                if car_plate_norm: new_op.car_from_api = car_plate_norm

                db.add(new_op)
                db.flush()
                new_count += 1
                # ОТПРАВКА ВОДИТЕЛЮ (если он найден и привязан к TG)
                if presumed_user and presumed_user.telegram_id:
                    try:
                        # Передаем bot, ID телеграма и ID созданной операции
                        await send_operation_to_user(message.bot, presumed_user.telegram_id, new_op.id)
                    except Exception as e:
                        logging.error(f"Ошибка отправки водителю {presumed_user.telegram_id}: {e}")

            db.commit()

        # СОБИРАЕМ АДМИНОВ В ПЛОСКИЙ СПИСОК (чтобы избежать DetachedInstanceError)
        if new_count > 0:
            recent_ops_raw = []
            with get_db_session() as db3:
                recent_ops = (
                    db3.query(
                        FuelOperation.id,
                        FuelOperation.doc_number,
                        FuelOperation.date_time,
                        FuelOperation.api_data,
                    )
                    .order_by(FuelOperation.imported_at.desc())
                    .limit(new_count)
                    .all()
                )
                for op_id, doc_number, date_time, api_data in recent_ops:
                    api = api_data if isinstance(api_data, dict) else {}
                    row_data = api.get("row", {})
                    card_data = api.get("card", {})

                    # Извлекаем данные (учитывая разные форматы API)
                    card_num = api.get("cardNumber") or api.get("card_number") or card_data.get("cardNumber")
                    azs = api.get("azsNumber") or api.get("azs") or row_data.get("azsNumber")
                    qty = api.get("productQuantity") or api.get("quantity") or row_data.get("productQuantity")
                    doc = doc_number or api.get("docNumber")
                    dt = date_time.strftime("%d.%m.%Y %H:%M") if date_time else "—"

                    recent_ops_raw.append({
                        "id": op_id, "doc": str(doc or "—"), "dt": str(dt),
                        "card": str(card_num or "—"), "azs": str(azs or "—"), "qty": str(qty or "—"),
                    })

            # Отправляем чеки ТОЛЬКО инициатору (вам)
            for op in recent_ops_raw:
                text = (
                    "⛽️ **Новая операция из API:**\n"
                    f"ID: `{op['id']}`\n"
                    f"Дата: {op['dt']}\n"
                    f"Карта: `{op['card']}`\n"
                    f"АЗС: {op['azs']}\n"
                    f"Чек: {op['doc']}\n"
                    f"Кол-во: {op['qty']} л."
                )
                kb = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"confirm_op:{op['id']}")],
                    [InlineKeyboardButton(text="👤 Назначить", callback_data=f"assign_op:{op['id']}")],
                    [InlineKeyboardButton(text="⚠️ Спорная", callback_data=f"mark_dispute:{op['id']}")],
                ])
                await message.answer(text, reply_markup=kb, parse_mode="Markdown")

        await message.reply(f"✅ Импорт завершён.\nНовых заправок в базе: {new_count}")

    except Exception as e:
        logger.exception("run_import_now")
        await message.reply(f"❌ Ошибка импорта: {e}")


@require_permission("admin:manage")
async def callback_confirm_op(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный ID.", show_alert=True)
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            await call.answer("Не найдено.", show_alert=True)
            return
        op.status = "confirmed"
        op.confirmed_at = datetime.now(timezone.utc)
        db.commit()
    try:
        await call.message.edit_reply_markup()
    except Exception:
        pass
    await call.message.answer(f"Операция {op_id} подтверждена.")
    await call.answer()


@require_permission("admin:manage")
async def callback_assign_op(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный ID.", show_alert=True)
        return
    await call.message.answer(f"Назначение: /assign_op {op_id} <user_id>")
    await call.answer()


@require_permission("admin:manage")
async def callback_mark_dispute(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный ID.", show_alert=True)
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            await call.answer("Не найдено.", show_alert=True)
            return
        op.status = "requires_manual"
        db.commit()
    try:
        await call.message.edit_reply_markup()
    except Exception:
        pass
    await call.message.answer(f"Операция {op_id} — ручная обработка.")
    await call.answer()


@require_permission("admin:manage")
async def cmd_assign_op(message: types.Message):
    args = extract_args(message).split()
    if len(args) < 2:
        await message.reply("Использование: /assign_op <op_id> <user_id>")
        return
    try:
        op_id = int(args[0])
        user_id = int(args[1])
    except ValueError:
        await message.reply("Числа: op_id и user_id.")
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        user = db.query(User).filter_by(id=user_id).first()
        if not op:
            await message.reply("Операция не найдена.")
            return
        if not user:
            await message.reply("Пользователь не найден.")
            return
        op.presumed_user_id = user.id
        db.commit()
        fn = user.full_name
    await message.reply(f"Операция {op_id} → пользователь {fn} (id={user_id}).")


@require_permission("admin:manage")
async def cmd_disputed_ops(message: types.Message):
    """Показать спорные операции"""
    with get_db_session() as db:
        ops = db.query(FuelOperation).filter(FuelOperation.status == "requires_manual").order_by(
            FuelOperation.id.desc()).limit(10).all()
        if not ops:
            await message.reply("✅ Нет спорных операций, требующих вмешательства.")
            return

        await message.reply("⚠️ **Спорные операции:**", parse_mode="Markdown")
        for op in ops:
            text = f"ID: {op.id} | Дата: {op.date_time}\nЧек: {op.doc_number} | Статус: Спорная"
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ Подтвердить вручную", callback_data=f"confirm_op:{op.id}")],
                [InlineKeyboardButton(text="👤 Назначить на...", callback_data=f"assign_op:{op.id}")]
            ])
            await message.answer(text, reply_markup=kb)


@require_permission("admin:manage")
async def cmd_recent_ops(message: types.Message):
    """Показать последние операции"""
    with get_db_session() as db:
        ops = db.query(FuelOperation).order_by(FuelOperation.id.desc()).limit(15).all()
        if not ops:
            await message.reply("База операций пуста.")
            return

        lines = [
            f"#{o.id} | {o.status} | {o.date_time.strftime('%d.%m %H:%M') if o.date_time else ''} | {o.car_from_api or ''}"
            for o in ops]
        await message.reply("🕒 **Последние 15 операций:**\n\n" + "\n".join(lines), parse_mode="Markdown")


@require_permission("admin:manage")
async def btn_update_import(message: types.Message):
    await cmd_run_import_now(message)


@require_permission("admin:manage")
async def btn_test_import(message: types.Message):
    await cmd_run_import_now_dry(message)


@require_permission("admin:manage")
async def btn_schedule_list(message: types.Message):
    await admin_schedules_mod.cmd_schedule_get(message)


@require_permission("admin:manage")
async def btn_schedule_set(message: types.Message):
    await message.reply("Команда: /schedule_set <имя> <HH:MM UTC>\nПример: /schedule_set daily 02:00")


@require_permission("admin:manage")
async def btn_schedule_remove(message: types.Message):
    await message.reply("Команда: /schedule_remove <имя>")


@require_permission("admin:manage")
async def btn_users(message: types.Message):
    await admin_users_mod.cmd_users(message)


@require_permission("admin:manage")
async def btn_generate_code(message: types.Message):
    await message.reply("Команда: /generate_code <user_id>")


@require_permission("admin:manage")
async def btn_export_codes(message: types.Message):
    await admin_users_mod.cmd_export_codes(message)


@require_permission("admin:manage")
async def btn_pending(message: types.Message):
    await cmd_pending_ops(message)


def register_admin_import_handlers(dp):
    dp.message.register(cmd_run_import_now, Command(commands=["run_import_now"]))
    dp.message.register(cmd_run_import_now_dry, Command(commands=["run_import_now_dry"]))
    dp.message.register(cmd_assign_op, Command(commands=["assign_op"]))
    dp.message.register(cmd_pending_ops, Command(commands=["pending_ops"]))
    dp.message.register(cmd_disputed_ops, F.text == BTN_ADMIN_DISPUTED)
    dp.message.register(cmd_recent_ops, F.text == BTN_ADMIN_RECENT)
    dp.message.register(btn_update_import, lambda m: m.text == BTN_ADMIN_IMPORT)
    dp.message.register(btn_test_import, lambda m: m.text == BTN_ADMIN_IMPORT_TEST)
    dp.message.register(btn_schedule_list, lambda m: m.text == BTN_ADMIN_SCHEDULES)
    dp.message.register(btn_schedule_set, lambda m: m.text == BTN_ADMIN_SCHEDULE_SET)
    dp.message.register(btn_schedule_remove, lambda m: m.text == BTN_ADMIN_SCHEDULE_DEL)
    dp.message.register(btn_users, lambda m: m.text == BTN_ADMIN_USERS)
    dp.message.register(btn_pending, lambda m: m.text == BTN_ADMIN_PENDING)
    dp.message.register(btn_export_excel, lambda m: m.text == BTN_ADMIN_EXPORT_EXCEL)
    dp.message.register(btn_generate_code, lambda m: m.text == BTN_ADMIN_GEN_CODE)
    dp.message.register(btn_export_codes, lambda m: m.text == BTN_ADMIN_EXPORT_CODES)
    dp.message.register(cmd_admin_help, lambda m: m.text == BTN_ADMIN_HELP)

    dp.callback_query.register(callback_confirm_op, lambda c: c.data and c.data.startswith("confirm_op:"))
    dp.callback_query.register(callback_assign_op, lambda c: c.data and c.data.startswith("assign_op:"))
    dp.callback_query.register(callback_mark_dispute, lambda c: c.data and c.data.startswith("mark_dispute:"))
