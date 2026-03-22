import os
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta
from io import StringIO, BytesIO
import csv

from aiogram import types, Dispatcher, F
from aiogram.types import (
    InputFile, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, BufferedInputFile
)
from aiogram.filters import Command
from sqlalchemy import cast, String, or_
from openpyxl import Workbook, load_workbook

from src.app.belorusneft_api import fetch_operational_raw, parse_operations
from src.app.config import TOKEN_SALT, CODE_TTL_HOURS
from src.app.tokens import verify_and_consume_code, generate_code, hash_code
from src.app.models import (
    LinkToken, User, FuelOperation, Schedule,
    FuelCard, Car, ConfirmationHistory, UserState # Добавлен UserState
)
from src.app.db import get_db_session
from src.app.permissions import require_permission, user_has_permission
from sqlalchemy.exc import IntegrityError

logger = logging.getLogger(__name__)

# Путь для сохранения Excel в корне проекта
EXPORT_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# В памяти: состояние подтверждений от пользователей
# telegram_id -> {"op_id": int, "step": "ask_car"|"ask_person", "attempts": int}
PENDING_OP_CONFIRM = {}

# --- Вспомогательные функции ---
def extract_args(message: types.Message) -> str:
    text = message.text or ""
    parts = text.split(maxsplit=1)
    return parts[1].strip() if len(parts) > 1 else ""


# --- Обработчики процесса подтверждения ---

async def callback_user_yes(call: types.CallbackQuery):
    """Обработка кнопки 'Да, это я'"""
    op_id = int(call.data.split(":")[1])
    with get_db_session() as db:
        user = db.query(User).filter_by(telegram_id=call.from_user.id).first()
        op = db.query(FuelOperation).filter_by(id=op_id).first()

        # Если в данных API уже есть госномер, заправка подтверждается сразу
        if op.car_from_api:
            op.status = "confirmed"
            op.confirmed_user_id = user.id
            op.confirmed_at = datetime.now(timezone.utc)
            db.commit()
            await call.message.edit_text(call.message.text + "\n\n✅ Подтверждено автоматически.")
            export_to_excel_final(op.id)
        else:
            # Если госномера нет, переходим в состояние ожидания ввода госномера
            db.merge(UserState(telegram_id=call.from_user.id, operation_id=op_id, step="ask_car"))
            db.commit()
            await call.message.answer("На каком автомобиле была заправка? Введите госномер (например, 1234 AB-7):")
    await call.answer()


async def callback_user_no(call: types.CallbackQuery):
    """Обработка кнопки 'Нет, не я' (Логика Пинг-понга)"""
    op_id = int(call.data.split(":")[1])
    with get_db_session() as db:
        user = db.query(User).filter_by(telegram_id=call.from_user.id).first()
        # Ищем, кто прислал этот запрос (для возврата)
        hist = db.query(ConfirmationHistory).filter_by(operation_id=op_id, to_user_id=user.id).order_by(
            ConfirmationHistory.id.desc()).first()

        if hist and hist.from_user_id:
            # Если задачу переслал другой юзер, возвращаем её ему
            prev_user = db.query(User).filter_by(id=hist.from_user_id).first()
            if prev_user and prev_user.telegram_id:
                await call.message.edit_text("❌ Задача возвращена инициатору.")
                await call.bot.send_message(
                    prev_user.telegram_id,
                    f"⚠️ Пользователь {user.full_name} отклонил заправку (ID {op_id}). Укажите верного заправщика (ФИО или госномер):"
                )
                db.merge(UserState(telegram_id=prev_user.telegram_id, operation_id=op_id, step="ask_person"))
        else:
            # Если это первый круг, просим текущего юзера указать, кто заправлялся
            db.merge(UserState(telegram_id=call.from_user.id, operation_id=op_id, step="ask_person"))
            await call.message.edit_text("❌ Укажите, кто заправлялся (ФИО или госномер):")

        db.commit()
    await call.answer()


async def handle_user_text(message: types.Message):
    """Обработка текстовых ответов на основе состояний (UserState)"""
    with get_db_session() as db:
        st = db.query(UserState).filter_by(telegram_id=message.from_user.id).first()
        if not st:
            return  # Если это просто текст, игнорируем

        op = db.query(FuelOperation).filter_by(id=st.operation_id).first()
        user = db.query(User).filter_by(telegram_id=message.from_user.id).first()
        text = message.text.strip()

        if st.step == "ask_car":
            op.actual_car = text.upper()
            op.status = "confirmed"
            op.confirmed_user_id = user.id
            op.confirmed_at = datetime.now(timezone.utc)
            db.delete(st)
            db.commit()
            await message.reply(f"✅ Госномер {text.upper()} сохранен. Заправка подтверждена.")
            export_to_excel_final(op.id)

        elif st.step == "ask_person":
            # Ищем человека по ФИО или госномерам его машин
            found = db.query(User).filter(User.full_name.ilike(f"%{text}%")).first()
            if not found:
                car = db.query(Car).filter(Car.plate.ilike(f"%{text}%")).first()
                if car and car.owners:
                    found = db.query(User).filter_by(id=car.owners[0]).first()

            if found and found.telegram_id:
                db.delete(st)
                db.commit()
                # Пересылаем операцию новому кандидату
                await send_operation_to_user(found.telegram_id, op.id, from_user_id=user.id)
                await message.reply(f"🚀 Запрос перенаправлен пользователю {found.full_name}.")
            else:
                await message.reply("Пользователь не найден. Попробуйте уточнить ФИО или введите госномер автомобиля:")


async def send_operation_to_user(tg_id: int, op_id: int, from_user_id: int = None):
    """Отправка запроса пользователю с полными данными из ТЗ"""
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        target_user = db.query(User).filter_by(telegram_id=tg_id).first()
        if not op or not target_user: return

        api = op.api_data or {}
        fuel = api.get("productName") or "—"
        qty = api.get("productQuantity") or "0"
        azs = api.get("azsNumber") or "—"
        dt = op.date_time.strftime("%d.%m.%Y %H:%M") if op.date_time else "—"

        text = (
            f"⛽️ *Подтвердите заправку*\n"
            f"📅 Дата: {dt}\n"
            f"⛽ Топливо: {fuel} ({qty} л.)\n"
            f"📍 АЗС: {azs}\n"
            f"💳 Карта: {api.get('cardNumber', '—')}\n"
            f"🧾 Чек: {op.doc_number or '—'}\n\n"
            f"Это ваша заправка?"
        )

        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да, это я", callback_data=f"user_yes:{op_id}")],
            [InlineKeyboardButton(text="❌ Нет, не я", callback_data=f"user_no:{op_id}")]
        ])

        # Фиксируем в истории, кому отправили
        db.add(ConfirmationHistory(operation_id=op_id, from_user_id=from_user_id, to_user_id=target_user.id, stage_result="sent"))
        db.commit()
        await types.Bot.get_current().send_message(tg_id, text, reply_markup=kb, parse_mode="Markdown")

# in-memory plain codes (shown once)
PENDING_PLAINS = {}  # token_id -> (plain_code, expires_at)


# --- User handlers ---
async def cmd_start(message: types.Message):
    kb_user = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="/link"), KeyboardButton(text="/myprofile")]],
        resize_keyboard=True
    )

    kb_admin = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 Обновить импорт"), KeyboardButton(text="🔎 Тестовый импорт")],
            [KeyboardButton(text="🗓 Расписания"), KeyboardButton(text="➕ Установить расписание")],
            [KeyboardButton(text="🗑 Удалить расписание"), KeyboardButton(text="👥 Пользователи")],
            [KeyboardButton(text="🔐 Сгенерировать код"), KeyboardButton(text="📤 Экспорт кодов")],
            [KeyboardButton(text="📊 Экспорт в Excel")]
        ],
        resize_keyboard=True
    )

    with get_db_session() as db:
        is_admin = user_has_permission(db, message.from_user.id, "admin:manage")

    if is_admin:
        await message.reply("Привет, админ. Выберите действие кнопкой или введите команду:", reply_markup=kb_admin)
    else:
        await message.reply("Привет! Для привязки аккаунта используйте /link <код>.", reply_markup=kb_user)


async def cmd_myprofile(message: types.Message):
    tg = message.from_user.id
    with get_db_session() as db:
        row = db.query(User.id, User.full_name, User.cards, User.cars).filter(User.telegram_id == tg).first()
    if not row:
        await message.reply("Профиль не найден. Привяжите аккаунт кодом /link <код>.")
        return
    _, full_name, cards, cars = row
    cards_s = ", ".join(cards or []) or "—"
    cars_s = ", ".join(cars or []) or "—"
    await message.reply(f"ФИО: {full_name}\nКарты: {cards_s}\nАвто: {cars_s}")


async def cmd_link(message: types.Message):
    args = extract_args(message)
    if not args:
        await message.reply("Введите код: /link <код>. Код вы получили от администратора.")
        return

    code = args.split()[0]
    with get_db_session() as db:
        ok, result = verify_and_consume_code(db, code, message.from_user.id)
        if not ok:
            reason = result
            user_row = None
        else:
            if isinstance(result, dict) and "user_id" in result:
                user_id = result["user_id"]
            elif hasattr(result, "user_id"):
                user_id = result.user_id
            else:
                user_id = None

            user_row = None
            if user_id:
                user_row = db.query(User.id, User.full_name, User.cards, User.cars).filter(User.id == user_id).first()

    if not ok:
        if reason == "invalid_or_used":
            await message.reply("Код неверен или уже использован. Обратитесь к администратору.")
        elif reason == "expired":
            await message.reply("Код просрочен. Попросите администратора сгенерировать новый.")
        elif reason == "already_linked_to_other":
            await message.reply("Эта запись уже привязана к другому Telegram. Обратитесь в поддержку.")
        else:
            await message.reply("Ошибка привязки. Обратитесь к администратору.")
        return

    if not user_row:
        await message.reply("Привязка выполнена, но не удалось получить данные пользователя.")
        return

    _, full_name, cards, cars = user_row
    cards_s = ", ".join(cards or []) or "—"
    cars_s = ", ".join(cars or []) or "—"
    info = f"Привязка выполнена.\nФИО: {full_name}\nКарты: {cards_s}\nАвто: {cars_s}"
    await message.reply(info)


# --- Admin: schedules ---
@require_permission("admin:manage")
async def cmd_schedule_get(message: types.Message):
    """
    Безопасно получить список расписаний и показать их.
    Извлекаем только скалярные поля внутри сессии, чтобы избежать DetachedInstanceError.
    """
    with get_db_session() as db:
        rows = db.query(
            Schedule.name,
            Schedule.cron_hour,
            Schedule.cron_minute,
            Schedule.enabled,
            Schedule.last_run
        ).order_by(Schedule.name).all()

        if not rows:
            await message.reply("Расписаний нет.")
            return

        lines = []
        for name, cron_hour, cron_minute, enabled, last_run in rows:
            last = last_run.isoformat() if last_run else "—"
            lines.append(f"{name}: {cron_hour:02d}:{cron_minute:02d} UTC — {'вкл' if enabled else 'выкл'} (last_run: {last})")

    # сессия закрыта, но мы уже сформировали строки
    await message.reply("\n".join(lines))



@require_permission("admin:manage")
async def cmd_schedule_set(message: types.Message):
    args = extract_args(message)
    parts = args.split()
    if len(parts) < 2:
        await message.reply("Использование: /schedule_set <name> <HH:MM UTC>")
        return
    name = parts[0]
    try:
        hh, mm = parts[1].split(":")
        hour = int(hh); minute = int(mm)
        if not (0 <= hour < 24 and 0 <= minute < 60):
            raise ValueError
    except Exception:
        await message.reply("Неверное время. Формат HH:MM (UTC).")
        return

    with get_db_session() as db:
        sched = db.query(Schedule).filter_by(name=name).first()
        if not sched:
            sched = Schedule(name=name, cron_hour=hour, cron_minute=minute, enabled=True)
            db.add(sched)
        else:
            sched.cron_hour = hour
            sched.cron_minute = minute
            sched.enabled = True
        db.commit()

    from src.app.scheduler import schedule_daily_import
    schedule_daily_import(name, hour, minute)
    await message.reply(f"Расписание {name} установлено на {hour:02d}:{minute:02d} UTC")


@require_permission("admin:manage")
async def cmd_schedule_remove(message: types.Message):
    args = extract_args(message).strip()
    if not args:
        await message.reply("Использование: /schedule_remove <name>")
        return
    name = args
    with get_db_session() as db:
        sched = db.query(Schedule).filter_by(name=name).first()
        if sched:
            db.delete(sched)
            db.commit()
    from src.app.scheduler import remove_schedule
    remove_schedule(name)
    await message.reply(f"Расписание {name} удалено.")


# --- Admin: users / tokens ---
@require_permission("admin:manage")
async def cmd_users(message: types.Message):
    args = extract_args(message)
    page = 1
    try:
        if args:
            page = max(1, int(args))
    except ValueError:
        page = 1

    page_size = 10
    offset = (page - 1) * page_size

    with get_db_session() as db:
        total = db.query(User).count()
        rows = db.query(User.id, User.full_name, User.telegram_id, User.active).order_by(User.id).offset(offset).limit(page_size).all()

    if not rows:
        await message.reply("Пользователи не найдены.")
        return

    for row in rows:
        user_id, full_name, telegram_id, active = row
        tg = f"@{telegram_id}" if telegram_id else "—"
        status = "Активен" if active else "Неактивен"
        text = f"{user_id}) {full_name} — {tg} — {status}"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="Сгенерировать код", callback_data=f"gen_code:{user_id}"),
                InlineKeyboardButton(text="Просмотр", callback_data=f"view_user:{user_id}")
            ]
        ])
        await message.answer(text, reply_markup=kb)

    pages = (total + page_size - 1) // page_size
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton(text="◀️ Назад", callback_data=f"users_page:{page-1}"))
    if page < pages:
        nav_buttons.append(InlineKeyboardButton(text="Вперёд ▶️", callback_data=f"users_page:{page+1}"))
    if nav_buttons:
        await message.answer(f"Страница {page}/{pages}", reply_markup=InlineKeyboardMarkup(inline_keyboard=[nav_buttons]))


@require_permission("admin:manage")
async def cmd_generate_code(message: types.Message):
    args = extract_args(message)
    if not args:
        await message.reply("Использование: /generate_code <user_id>")
        return
    try:
        user_id = int(args.split()[0])
    except ValueError:
        await message.reply("Неверный user_id. Использование: /generate_code <user_id>")
        return

    admin_id = message.from_user.id
    with get_db_session() as db:
        user_row = db.query(User.id, User.full_name).filter_by(id=user_id).first()
        if not user_row:
            await message.reply(f"Пользователь с id={user_id} не найден.")
            return
        _, full_name = user_row

        code_plain = generate_code()
        code_hash = hash_code(code_plain, TOKEN_SALT)
        expires_at = datetime.now(timezone.utc) + timedelta(hours=CODE_TTL_HOURS)

        token = LinkToken(user_id=user_id, code_hash=code_hash, created_by=admin_id, created_at=datetime.now(timezone.utc), expires_at=expires_at)
        try:
            db.add(token)
            db.flush()
            token_id = token.id
            db.commit()
        except IntegrityError:
            db.rollback()
            await message.reply("Не удалось сгенерировать код (конфликт). Попробуйте ещё раз.")
            return

    PENDING_PLAINS[token_id] = (code_plain, datetime.now(timezone.utc) + timedelta(minutes=10))

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Отправить пользователю", callback_data=f"send_code:{token_id}")],
        [InlineKeyboardButton(text="Отозвать код", callback_data=f"revoke_code:{token_id}")]
    ])
    await message.reply(f"Код для пользователя {full_name} (id={user_id}):\n\n{code_plain}\n\nКод показывается один раз.", reply_markup=kb)


@require_permission("admin:manage")
async def cmd_export_codes(message: types.Message):
    args_list = extract_args(message).split()
    user_id = None
    if args_list and args_list[0]:
        try:
            user_id = int(args_list[0])
        except ValueError:
            await message.reply("Неверный user_id. Использование: /export_codes [user_id]")
            return

    with get_db_session() as db:
        query = db.query(LinkToken)
        if user_id:
            query = query.filter(LinkToken.user_id == user_id)
        tokens = query.all()

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "token_id", "user_id", "user_full_name", "code_hash", "created_by",
            "created_at", "expires_at", "status", "telegram_id", "used_at", "note"
        ])
        for t in tokens:
            user = db.query(User).filter_by(id=t.user_id).first() if t.user_id else None
            writer.writerow([
                t.id,
                t.user_id or "",
                user.full_name if user else "",
                getattr(t, "code_hash", "") or "",
                t.created_by or "",
                t.created_at.isoformat() if t.created_at else "",
                t.expires_at.isoformat() if t.expires_at else "",
                getattr(t, "status", "") or "",
                t.telegram_id or "",
                t.used_at.isoformat() if getattr(t, "used_at", None) else "",
                t.note or ""
            ])
        csv_bytes = output.getvalue().encode("utf-8")
        output.close()

        bio = BytesIO(csv_bytes)
        bio.name = f"link_tokens_{user_id or 'all'}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.csv"
        bio.seek(0)
        await message.reply_document(InputFile(bio, filename=bio.name))


# --- Callbacks for user list / tokens ---
def export_to_excel_final(op_id: int):
    """Создает запись в Excel по 23 колонкам из ТЗ"""
    file_path = EXPORT_DIR / "Fuel_Report_Master.xlsx"
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op: return
        api = op.api_data or {}
        presumed = db.query(User).filter_by(id=op.presumed_user_id).first()
        confirmed = db.query(User).filter_by(id=op.confirmed_user_id).first()

        row = [
            op.id, "Топливная карта" if op.source == "api" else "Личные средства",
            op.source, op.date_time.strftime("%d.%m.%Y") if op.date_time else "",
            op.date_time.strftime("%H:%M:%S") if op.date_time else "",
            api.get("cardNumber") or "—", api.get("productName") or "—",
            api.get("productQuantity") or 0, api.get("productCost") or 0,
            api.get("azsNumber") or "—", op.doc_number or "—",
            op.car_from_api or "—", api.get("driverName") or "—", "",
            presumed.full_name if presumed else "—", confirmed.full_name if confirmed else "—",
            op.actual_car or op.car_from_api or "—", "Система",
            confirmed.full_name if confirmed else "—", op.status,
            op.confirmed_at.strftime("%d.%m.%Y %H:%M") if op.confirmed_at else "—",
            "", "Да" if op.status == "confirmed" else "Нет"
        ]
        try:
            if not file_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["ID", "Тип", "Источник", "Дата", "Время", "Карта", "Топливо", "Литры", "Сумма", "АЗС", "Чек",
                           "Авто(API)", "Водитель(API)", "OCR", "Предполагаемый", "Фактический", "Факт.Авто",
                           "Инициатор", "Подтвердил", "Статус", "Дата Подтв", "Прим", "Готов"])
            else:
                wb = load_workbook(file_path)
                ws = wb.active
            ws.append(row)
            wb.save(file_path)
            op.exported_to_excel = True
            db.commit()
        except Exception as e:
            logger.error(f"Excel error: {e}")


async def send_operation_to_user(tg_id: int, op_id: int, from_user_id: int = None):
    """Отправка сообщения с кнопками Да/Нет"""
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op: return
        api = op.api_data or {}
        text = (f"⛽️ *Подтвердите заправку*\n\nДата: {op.date_time}\nТопливо: {api.get('productName')}\n"
                f"Литры: {api.get('productQuantity')}\nЧек: {op.doc_number}\n\nЭто вы заправлялись?")
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да", callback_data=f"user_yes:{op_id}"),
             InlineKeyboardButton(text="❌ Нет", callback_data=f"user_no:{op_id}")]
        ])
        db.add(ConfirmationHistory(operation_id=op_id, from_user_id=from_user_id,
                                   to_user_id=db.query(User).filter_by(telegram_id=tg_id).first().id,
                                   stage_result="sent"))
        db.commit()
        await types.Bot.get_current().send_message(tg_id, text, reply_markup=kb, parse_mode="Markdown")



def export_operation_to_excel(op_id: int):
    """
    Экспорт одной подтверждённой операции в Excel.
    Сохраняет файл exports/operations_YYYYMMDD.xlsx и помечает запись.
    """
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            raise RuntimeError("Operation not found")

        # собираем данные
        api = op.api_data or {}
        row = {
            "id": op.id,
            "date_time": getattr(op, "date_time", None) or api.get("dateTimeIssue"),
            "card": api.get("cardNumber") or api.get("card_number") or "—",
            "azs": api.get("azsNumber") or api.get("azs") or "—",
            "doc": getattr(op, "doc_number", None) or api.get("docNumber") or api.get("doc_number") or "—",
            "quantity": api.get("productQuantity") or api.get("quantity") or "—",
            "presumed_user_id": op.presumed_user_id or "",
            "car": op.car_from_api or "",
            "status": op.status or "",
        }

        # файл по дате
        date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
        path = EXPORT_DIR / f"operations_{date_str}.xlsx"

        # если файл существует — открываем, иначе создаём
        if path.exists():
            wb = Workbook()
            wb = wb.load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = ["id", "date_time", "card", "azs", "doc", "quantity", "presumed_user_id", "car", "status"]
            ws.append(headers)

        ws.append([row[h] for h in ["id", "date_time", "card", "azs", "doc", "quantity", "presumed_user_id", "car", "status"]])
        wb.save(path)

        # пометка в БД
        op.exported_to_excel = True
        db.commit()





def export_to_excel_final(op_id: int):
    """Экспорт согласно Приложению 1 (23 колонки)"""
    file_path = EXPORT_DIR / "Fuel_Report_Master.xlsx"
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        api = op.api_data or {}

        # Подготовка данных (23 колонки по ТЗ)
        row = [
            op.id, "Топливная карта", "API",
            op.date_time.date() if op.date_time else "",
            op.date_time.time() if op.date_time else "",
            api.get("cardNumber"), api.get("productName"), api.get("productQuantity"),
            api.get("productCost"), api.get("azsNumber"), op.doc_number,
            op.car_from_api, api.get("driverName"), "",  # OCR пустой
            "Предполагаемый", "Фактический", op.actual_car, "Система",
            "Кто подтвердил", op.status, datetime.now().strftime("%d.%m.%Y"), "", "Да"
        ]

        if not file_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Заправки по картам"
            ws.append(["ID", "Тип", "Источник", "Дата", "Время", "Карта", "Топливо", "Кол-во", "Сумма", "АЗС", "Чек",
                       "Авто(API)", "Водитель(API)", "OCR", "Предполагаемый", "Фактический", "Факт.Авто", "Инициатор",
                       "Подтвердил", "Статус", "Дата подтв", "Прим", "Готовность"])
        else:
            wb = load_workbook(file_path)
            ws = wb["Заправки по картам"]

        ws.append(row)
        wb.save(file_path)
async def callback_users_page(call: types.CallbackQuery):
    page = int(call.data.split(":", 1)[1])
    await call.message.delete()
    fake_msg = types.Message(**{
        "message_id": call.message.message_id,
        "date": call.message.date,
        "chat": call.message.chat,
        "from_user": call.from_user,
        "text": f"/users {page}"
    })
    await cmd_users(fake_msg)
    await call.answer()


async def callback_view_user(call: types.CallbackQuery):
    user_id = int(call.data.split(":", 1)[1])
    with get_db_session() as db:
        row = db.query(User.id, User.full_name, User.telegram_id, User.cards, User.cars, User.active).filter(User.id == user_id).first()
    if not row:
        await call.message.answer("Пользователь не найден.")
        await call.answer()
        return
    uid, full_name, telegram_id, cards, cars, active = row
    tg = f"@{telegram_id}" if telegram_id else "—"
    cards_s = ", ".join(cards or []) or "—"
    cars_s = ", ".join(cars or []) or "—"
    text = (
        f"ID: {uid}\n"
        f"ФИО: {full_name}\n"
        f"Telegram: {tg}\n"
        f"Карты: {cards_s}\n"
        f"Авто: {cars_s}\n"
        f"Активен: {'Да' if active else 'Нет'}\n"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Сгенерировать код", callback_data=f"gen_code:{uid}")],
        [InlineKeyboardButton(text="Закрыть", callback_data="noop")]
    ])
    await call.message.answer(text, reply_markup=kb)
    await call.answer()


async def callback_generate_code(call: types.CallbackQuery):
    user_id = int(call.data.split(":", 1)[1])
    admin_id = call.from_user.id
    with get_db_session() as db:
        row = db.query(User.id, User.full_name, User.telegram_id).filter(User.id == user_id).first()
        if not row:
            await call.message.answer("Пользователь не найден.")
            await call.answer()
            return
        uid, full_name, telegram_id = row

        code_plain = generate_code()
        code_hash = hash_code(code_plain, TOKEN_SALT)
        expires_at = datetime.now(timezone.utc) + timedelta(hours=CODE_TTL_HOURS)
        token = LinkToken(user_id=uid, code_hash=code_hash, created_by=admin_id, created_at=datetime.now(timezone.utc), expires_at=expires_at)
        try:
            db.add(token)
            db.flush()
            token_id = token.id
            db.commit()
        except IntegrityError:
            db.rollback()
            await call.message.answer("Не удалось сгенерировать код (конфликт). Попробуйте ещё раз.")
            await call.answer()
            return

    PENDING_PLAINS[token_id] = (code_plain, datetime.now(timezone.utc) + timedelta(minutes=10))

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Отправить пользователю", callback_data=f"send_code:{token_id}")],
        [InlineKeyboardButton(text="Отозвать код", callback_data=f"revoke_code:{token_id}")]
    ])
    await call.message.answer(f"Код для {full_name} (id={uid}):\n\n{code_plain}\n\nКод показывается один раз.", reply_markup=kb)
    await call.answer()


async def callback_send_code(call: types.CallbackQuery):
    token_id = int(call.data.split(":", 1)[1])

    entry = PENDING_PLAINS.get(token_id)
    if entry and entry[1] <= datetime.now(timezone.utc):
        del PENDING_PLAINS[token_id]
        entry = None
    plain_code = entry[0] if entry else None

    with get_db_session() as db:
        token = db.query(LinkToken).filter_by(id=token_id).first()
        if not token:
            await call.message.answer("Токен не найден.")
            await call.answer()
            return
        user_row = db.query(User.id, User.full_name, User.telegram_id).filter(User.id == token.user_id).first()

    if user_row and user_row[2]:
        tg_id = user_row[2]
        if plain_code:
            try:
                await call.bot.send_message(tg_id, f"Вам выдан код для привязки аккаунта: {plain_code}\nВведите его в боте: /link <код>")
                await call.message.answer("Код отправлен пользователю.")
            except Exception:
                await call.message.answer("Не удалось отправить сообщение пользователю (возможно, пользователь не начинал диалог с ботом).")
        else:
            await call.message.answer("Plain-код недоступен (время истекло). Передайте код вручную или сгенерируйте новый.")
    else:
        if user_row:
            await call.message.answer(f"У пользователя {user_row[1]} нет привязанного Telegram. Передайте ему код вручную.")
        else:
            await call.message.answer("Пользователь не найден. Невозможно отправить код.")
    await call.answer()


async def callback_revoke_code(call: types.CallbackQuery):
    token_id = int(call.data.split(":", 1)[1])
    admin_id = call.from_user.id
    with get_db_session() as db:
        token = db.query(LinkToken).filter_by(id=token_id).first()
        if not token:
            await call.message.answer("Токен не найден.")
            await call.answer()
            return
        token.status = "revoked"
        token.note = (token.note or "") + f"\nRevoked by admin {admin_id} at {datetime.now(timezone.utc).isoformat()}"
        db.commit()
    await call.message.answer("Код отозван.")
    await call.answer()


# --- Import operations (admin) ---
@require_permission("admin:manage")
async def cmd_run_import_now_dry(message: types.Message):
    from src.app.jobs import run_import_job
    try:
        run_import_job("manual_dry_run", dry_run=True)
        await message.reply("Тестовый импорт выполнен (dry‑run). Проверьте логи и debug‑дампы. Ничего не сохранено.")
    except Exception as e:
        logging.getLogger(__name__).exception("Error during dry-run import")
        await message.reply(f"Ошибка при выполнении тестового импорта: {e}")


@require_permission("admin:manage")
async def cmd_run_import_now(message: types.Message):
    new_count = 0
    date = datetime.now() - timedelta(days=1)
    try:
        raw = fetch_operational_raw(date)
        status = raw.get("status")
        json_payload = raw.get("json")
        debug_files = raw.get("debug_files") or []

        if status != 200:
            await message.reply(f"Запрос вернул статус {status}. Debug файлы: {', '.join(debug_files)}")
            return

        if not json_payload:
            await message.reply(f"JSON не получен. Debug файлы: {', '.join(debug_files)}")
            return

        ops = parse_operations(json_payload)
        if not ops:
            await message.reply("В ответе найдено элементов: 0 (импорт временно отключён)")
            return


        with get_db_session() as db:
            for op in ops:
                # parse date
                dt_raw = op.get("date_time") or op.get("dateTimeIssue")
                dt_obj = None
                if dt_raw:
                    try:
                        dt_obj = datetime.fromisoformat(dt_raw)
                    except Exception:
                        dt_obj = None

                # normalize doc
                doc_raw = op.get("doc_number") or op.get("docNumber")
                doc = None
                if doc_raw is not None:
                    try:
                        doc = str(doc_raw).strip()
                    except Exception:
                        doc = None

                # ← ВОТ СЮДА ВСТАВЛЯЕМ
                logging.info(f"Checking op: doc={doc}, dt={dt_obj}")

                # dedupe
                filters = [FuelOperation.source == "api"]
                if doc and hasattr(FuelOperation, "doc_number"):
                    filters.append(FuelOperation.doc_number == doc)
                if dt_obj and hasattr(FuelOperation, "date_time"):
                    filters.append(FuelOperation.date_time == dt_obj)

                exists = db.query(FuelOperation).filter(*filters).first()
                if exists:
                    continue

                # FuelCard
                card_num = op.get("card_number") or op.get("cardNumber") or op.get("card")
                if card_num is not None:
                    try:
                        card_num = str(card_num).strip()
                    except Exception:
                        card_num = None

                fuel_card = None
                if card_num:
                    fuel_card = db.query(FuelCard).filter_by(card_number=card_num).first()
                    if not fuel_card:
                        fuel_card = FuelCard(card_number=card_num, active=True)
                        db.add(fuel_card)
                        db.flush()

                # presumed user
                presumed_user = None
                if fuel_card and fuel_card.user_id:
                    presumed_user = db.query(User).filter_by(id=fuel_card.user_id).first()
                else:
                    if card_num:
                        try:
                            presumed_user = (
                                db.query(User)
                                .filter(cast(User.cards, String).like(f"%{card_num}%"))
                                .first()
                            )
                        except Exception:
                            presumed_user = None

                        if presumed_user and fuel_card:
                            fuel_card.user_id = presumed_user.id

                # Car handling
                car_plate = op.get("carNum") or op.get("car_num") or op.get("car")
                car_plate_norm = None
                if car_plate:
                    car_plate_norm = str(car_plate).strip().upper()
                    car = db.query(Car).filter_by(plate=car_plate_norm).first()
                    if not car:
                        car = Car(plate=car_plate_norm)
                        db.add(car)
                        db.flush()
                    if presumed_user:
                        owners = car.owners or []
                        if presumed_user.id not in owners:
                            owners.append(presumed_user.id)
                            car.owners = owners
                        user_cars = presumed_user.cars or []
                        if car_plate_norm not in user_cars:
                            user_cars.append(car_plate_norm)
                            presumed_user.cars = user_cars

                # create operation
                new_op = FuelOperation(
                    source="api",
                    api_data=op.get("raw") or op,
                    imported_at=datetime.now(timezone.utc),
                    status="loaded"
                )
                if hasattr(FuelOperation, "doc_number") and doc:
                    new_op.doc_number = doc
                if hasattr(FuelOperation, "date_time") and dt_obj:
                    new_op.date_time = dt_obj
                if presumed_user:
                    new_op.presumed_user_id = presumed_user.id
                if car_plate_norm:
                    new_op.car_from_api = car_plate_norm

                db.add(new_op)
                db.flush()
                new_count += 1

                # если есть предполагаемый пользователь с telegram_id — запланируем уведомление
                if presumed_user and getattr(presumed_user, "telegram_id", None):
                    # собираем пары для уведомления после коммита
                    if "pending_notifications" not in locals():
                        pending_notifications = []
                    pending_notifications.append((presumed_user.telegram_id, new_op.id))

            db.commit()

            # после коммита отправляем пользователям запросы на подтверждение
            pending_notifications = locals().get("pending_notifications", []) or []
            for tg_id, op_id in pending_notifications:
                try:
                    await send_operation_to_user(tg_id, op_id)
                except Exception:
                    logging.getLogger(__name__).exception("Failed to send operation %s to user %s", op_id, tg_id)

        # notify admins — SAFE VERSION (no DetachedInstanceError)
        # Собираем список админов как простые dict внутри сессии (чтобы избежать detached instances)
        admin_rows = []
        with get_db_session() as db2:
            admins_q = db2.query(User.id, User.telegram_id, User.full_name).filter(User.telegram_id != None).all()
            for uid, tg_id, full_name in admins_q:
                try:
                    if user_has_permission(db2, tg_id, "admin:manage"):
                        # сохраняем только скалярные поля
                        admin_rows.append({"id": uid, "telegram_id": tg_id, "full_name": full_name})
                except Exception:
                    continue

        if new_count > 0 and admin_rows:
            # Собираем данные операций заранее, пока сессия открыта
            with get_db_session() as db3:
                recent_ops_raw = []
                recent_ops = (
                    db3.query(
                        FuelOperation.id,
                        FuelOperation.doc_number,
                        FuelOperation.date_time,
                        FuelOperation.api_data,
                    )
                    .order_by(FuelOperation.imported_at.desc())
                    .limit(new_count)
                    .all()
                )

                for op_id, doc_number, date_time, api_data in recent_ops:
                    api = api_data or {}
                    recent_ops_raw.append({
                        "id": op_id,
                        "doc": doc_number or api.get("docNumber") or api.get("doc_number"),
                        "dt": date_time.isoformat() if date_time else api.get("dateTimeIssue"),
                        "card": api.get("cardNumber") or api.get("card_number") or "—",
                        "azs": api.get("azsNumber") or api.get("azs") or "—",
                        "qty": api.get("productQuantity") or api.get("quantity") or "—",
                    })

            # Теперь отправляем уведомления, работая только с простыми dict
            for admin in admin_rows:
                tg_id = admin.get("telegram_id")
                if not tg_id:
                    continue
                for op in recent_ops_raw:
                    text = (
                        "Новая операция из API:\n"
                        f"ID: {op['id']}\n"
                        f"Дата: {op['dt']}\n"
                        f"Карта: {op['card']}\n"
                        f"АЗС: {op['azs']}\n"
                        f"Чек: {op['doc']}\n"
                        f"Кол-во: {op['qty']}"
                    )
                    kb = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"confirm_op:{op['id']}")],
                        [InlineKeyboardButton(text="👤 Назначить пользователя", callback_data=f"assign_op:{op['id']}")],
                        [InlineKeyboardButton(text="⚠️ Пометить спорной", callback_data=f"mark_dispute:{op['id']}")]
                    ])
                    try:
                        await message.bot.send_message(tg_id, text, reply_markup=kb)
                    except Exception:
                        logging.getLogger(__name__).warning(
                            "Не удалось отправить администратору %s сообщение о операции %s",
                            admin.get("id"), op["id"]
                        )

        await message.reply(f"Импорт завершён. Новых операций: {new_count}")

    except Exception as e:
        logging.getLogger(__name__).exception("Error during run_import_now")
        await message.reply(f"Ошибка при выполнении запроса: {e}")

        def format_op_short(op):
            api = op.api_data or {}
            doc = getattr(op, "doc_number", "") or api.get("docNumber", "") or api.get("doc_number", "")
            dt = getattr(op, "date_time", "") or api.get("dateTimeIssue", "") or api.get("date_time", "")
            card = api.get("cardNumber") or api.get("card_number") or "—"
            azs = api.get("azsNumber") or api.get("azs") or "—"
            qty = api.get("productQuantity") or api.get("quantity") or "—"
            return f"ID:{op.id}\nДата: {dt}\nКарта: {card}\nАЗС: {azs}\nЧек: {doc}\nКол-во: {qty}"

        admin_rows = []
        with get_db_session() as db2:
            for u in db2.query(User).filter(User.telegram_id != None).all():
                try:
                    if user_has_permission(db2, u.telegram_id, "admin:manage"):
                        admin_rows.append(u)
                except Exception:
                    continue

        if new_count > 0 and admin_rows:
            with get_db_session() as db3:
                recent_ops = db3.query(FuelOperation).order_by(FuelOperation.imported_at.desc()).limit(new_count).all()
            for admin in admin_rows:
                for op in recent_ops:
                    text = "Новая операция из API:\n" + format_op_short(op)
                    kb = InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"confirm_op:{op.id}")],
                        [InlineKeyboardButton(text="👤 Назначить пользователя", callback_data=f"assign_op:{op.id}")],
                        [InlineKeyboardButton(text="⚠️ Пометить спорной", callback_data=f"mark_dispute:{op.id}")]
                    ])
                    try:
                        await message.bot.send_message(admin.telegram_id, text, reply_markup=kb)
                    except Exception:
                        logging.getLogger(__name__).warning(
                            "Не удалось отправить администратору %s сообщение о операции %s", admin.id, op.id)

        await message.reply(f"Импорт завершён. Новых операций: {new_count}")

    except Exception as e:
        logging.getLogger(__name__).exception("Error during run_import_now")
        await message.reply(f"Ошибка при выполнении запроса: {e}")


# --- Callbacks for admin actions on imported operations ---
@require_permission("admin:manage")
async def callback_confirm_op(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный идентификатор операции.", show_alert=True)
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            await call.answer("Операция не найдена.", show_alert=True)
            return
        op.status = "confirmed"
        op.confirmed_at = datetime.now(timezone.utc)
        db.commit()
    try:
        await call.message.edit_reply_markup()
    except Exception:
        pass
    await call.message.answer(f"Операция {op_id} помечена как подтверждённая.")
    await call.answer()


@require_permission("admin:manage")
async def callback_assign_op(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный идентификатор операции.", show_alert=True)
        return
    await call.message.answer(f"Чтобы назначить операцию {op_id} пользователю, выполните команду:\n/assign_op {op_id} <user_id>")
    await call.answer()


@require_permission("admin:manage")
async def callback_mark_dispute(call: types.CallbackQuery):
    try:
        op_id = int(call.data.split(":", 1)[1])
    except Exception:
        await call.answer("Неверный идентификатор операции.", show_alert=True)
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            await call.answer("Операция не найдена.", show_alert=True)
            return
        op.status = "requires_manual"
        db.commit()
    try:
        await call.message.edit_reply_markup()
    except Exception:
        pass
    await call.message.answer(f"Операция {op_id} помечена как спорная и переведена в ручную обработку.")
    await call.answer()


@require_permission("admin:manage")
async def cmd_assign_op(message: types.Message):
    args = extract_args(message).split()
    if len(args) < 2:
        await message.reply("Использование: /assign_op <op_id> <user_id>")
        return
    try:
        op_id = int(args[0]); user_id = int(args[1])
    except ValueError:
        await message.reply("op_id и user_id должны быть числами.")
        return
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        user = db.query(User).filter_by(id=user_id).first()
        if not op:
            await message.reply("Операция не найдена.")
            return
        if not user:
            await message.reply("Пользователь не найден.")
            return
        op.presumed_user_id = user.id
        db.commit()
    await message.reply(f"Операция {op_id} назначена пользователю {user.full_name} (id={user.id}).")


# --- ReplyKeyboard wrappers (text buttons) ---
@require_permission("admin:manage")
async def btn_update_import(message: types.Message):
    await cmd_run_import_now(message)

@require_permission("admin:manage")
async def btn_test_import(message: types.Message):
    await cmd_run_import_now_dry(message)

@require_permission("admin:manage")
async def btn_schedule_list(message: types.Message):
    await cmd_schedule_get(message)

@require_permission("admin:manage")
async def btn_schedule_set(message: types.Message):
    await message.reply("Использование: /schedule_set <name> <HH:MM UTC>\nПример: /schedule_set belorusneft_daily 01:30 UTC")

@require_permission("admin:manage")
async def btn_schedule_remove(message: types.Message):
    await message.reply("Использование: /schedule_remove <name>\nПример: /schedule_remove belorusneft_daily")

@require_permission("admin:manage")
async def btn_users(message: types.Message):
    await cmd_users(message)

@require_permission("admin:manage")
async def btn_generate_code(message: types.Message):
    await message.reply("Использование: /generate_code <user_id>\nПример: /generate_code 42")

@require_permission("admin:manage")
async def btn_export_codes(message: types.Message):
    await cmd_export_codes(message)


@require_permission("admin:manage")
async def btn_export_excel(message: types.Message):
    """Генерация и отправка файла Excel со всеми заправками"""
    await message.answer("⏳ Собираю данные для отчета, подождите...")

    with get_db_session() as db:
        operations = db.query(FuelOperation).order_by(FuelOperation.date_time.desc()).all()

        if not operations:
            await message.answer("Нет данных для экспорта.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Заправки"

        # Заголовки (согласно вашему ТЗ)
        headers = [
            "ID", "Дата", "Время", "Источник", "Статус",
            "Карта", "Топливо", "Объем (л)", "Стоимость",
            "АЗС", "Госномер (из API)", "Водитель (из API)",
            "Фактический Госномер", "Подтвердил (ФИО)"
        ]
        ws.append(headers)

        for op in operations:
            api = op.api_data or {}
            # Собираем строку данных
            ws.append([
                op.id,
                op.date_time.strftime('%d.%m.%Y') if op.date_time else "—",
                op.date_time.strftime('%H:%M') if op.date_time else "—",
                op.source,
                op.status,
                api.get('cardNumber', '—'),
                api.get('productName', '—'),
                api.get('productQuantity', 0),
                api.get('productCost', 0),
                api.get('azsNumber', '—'),
                api.get('carNum', '—'),
                api.get('driverName', '—'),
                op.actual_car or "—",
                op.confirmed_user.full_name if op.confirmed_user else "—"
            ])

        # Сохраняем в буфер, чтобы не мусорить файлами на диске
        file_buffer = BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        filename = f"Fuel_Report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

        await message.answer_document(
            types.BufferedInputFile(file_buffer.read(), filename=filename),
            caption=f"📊 Выгрузка заправок на {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )

# --- Register handlers ---
def register_handlers(dp: Dispatcher):
    # commands
    dp.message.register(cmd_start, Command(commands=["start"]))
    dp.message.register(cmd_link, Command(commands=["link"]))
    dp.message.register(cmd_myprofile, Command(commands=["myprofile"]))
    dp.message.register(cmd_users, Command(commands=["users"]))
    dp.message.register(cmd_generate_code, Command(commands=["generate_code"]))
    dp.message.register(cmd_export_codes, Command(commands=["export_codes"]))
    dp.message.register(cmd_run_import_now, Command(commands=["run_import_now"]))
    dp.message.register(cmd_run_import_now_dry, Command(commands=["run_import_now_dry"]))
    dp.message.register(cmd_schedule_get, Command(commands=["schedule_get"]))
    dp.message.register(cmd_schedule_set, Command(commands=["schedule_set"]))
    dp.message.register(cmd_schedule_remove, Command(commands=["schedule_remove"]))
    dp.message.register(cmd_assign_op, Command(commands=["assign_op"]))

    # text buttons -> handlers
    dp.message.register(btn_update_import, lambda m: m.text == "📥 Обновить импорт")
    dp.message.register(btn_test_import, lambda m: m.text == "🔎 Тестовый импорт")
    dp.message.register(btn_schedule_list, lambda m: m.text == "🗓 Расписания")
    dp.message.register(btn_schedule_set, lambda m: m.text == "➕ Установить расписание")
    dp.message.register(btn_schedule_remove, lambda m: m.text == "🗑 Удалить расписание")
    dp.message.register(btn_users, lambda m: m.text == "👥 Пользователи")
    dp.message.register(btn_generate_code, lambda m: m.text == "🔐 Сгенерировать код")
    dp.message.register(btn_export_codes, lambda m: m.text == "📤 Экспорт кодов")

    # callback handlers
    dp.callback_query.register(callback_users_page, lambda c: c.data and c.data.startswith("users_page:"))
    dp.callback_query.register(callback_view_user, lambda c: c.data and c.data.startswith("view_user:"))
    dp.callback_query.register(callback_generate_code, lambda c: c.data and c.data.startswith("gen_code:"))
    dp.callback_query.register(callback_send_code, lambda c: c.data and c.data.startswith("send_code:"))
    dp.callback_query.register(callback_revoke_code, lambda c: c.data and c.data.startswith("revoke_code:"))
    dp.callback_query.register(lambda c: c.answer(), lambda c: c.data == "noop")
    dp.message.register(btn_export_excel, F.text == "📊 Экспорт в Excel")  # Регистрация кнопки
    # Пример регистрации в основном модуле, где есть Dispatcher dp
    dp.callback_query.register(callback_user_yes, F.data.startswith("user_yes:"))
    dp.callback_query.register(callback_user_no, F.data.startswith("user_no:"))
    dp.message.register(handle_user_text, F.text)
    dp.message.register(handle_user_text, F.text & ~F.text.startswith("/"))
    # admin callbacks for imported operations
    dp.callback_query.register(callback_confirm_op, lambda c: c.data and c.data.startswith("confirm_op:"))
    dp.callback_query.register(callback_assign_op, lambda c: c.data and c.data.startswith("assign_op:"))
    dp.callback_query.register(callback_mark_dispute, lambda c: c.data and c.data.startswith("mark_dispute:"))
