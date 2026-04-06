"""
Выгрузка подтверждённых и спорных операций в Excel (структура по ТЗ).
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.app.db import get_db_session
from src.app.models import FuelOperation, User, ConfirmationHistory

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent.parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
MASTER_FILE = EXPORT_DIR / "Fuel_Report_Master.xlsx"

SHEET_CARDS = "Заправки_по_картам"
SHEET_DISPUTED = "Спорные заправки"
SHEET_PERSONAL = "Заправки_личные_средства"
SHEET_REF = "Справочники"

STATUS_RU = {
    "loaded_from_api": "Загружена из API",
    "pending": "Ожидает подтверждения",
    "confirmed": "Подтверждена",
    "requires_manual": "Требует ручной обработки",
    "disputed": "Спорная",
    "rejected_by_other": "Отклонена",
    "import_error": "Ошибка импорта"
}

HEADERS = [
    "ID",
    "Тип заправки",
    "Источник",
    "Дата",
    "Время",
    "Карта",
    "Топливо",
    "Литры",
    "Сумма",
    "АЗС",
    "Чек",
    "Авто(API)",
    "Водитель(API)",
    "Данные OCR",
    "Предполагаемый пользователь",
    "Фактически подтвердивший",
    "Фактическое авто",
    "Кто первоначально получил запрос",
    "Кто окончательно подтвердил",
    "Статус подтверждения",
    "Дата и время подтверждения",
    "Примечание",
    "Готовность к путевому листу"
]


def _first_confirmation_sender_name(db, op_id: int) -> str:
    """Кто первоначально получил запрос (для чеков — отправивший фото)."""
    op = db.query(FuelOperation).filter_by(id=op_id).first()
    if op and op.source == "personal_receipt" and op.presumed_user_id:
        u = db.query(User).filter_by(id=op.presumed_user_id).first()
        return u.full_name if u else "—"
    h = (
        db.query(ConfirmationHistory)
        .filter_by(operation_id=op_id)
        .order_by(ConfirmationHistory.id.asc())
        .first()
    )
    if not h or not getattr(h, "to_user_id", None):
        return "—"
    u = db.query(User).filter_by(id=h.to_user_id).first()
    return u.full_name if u else "—"


def _ocr_text(op: FuelOperation) -> str:
    """Колонка OCR: сырой текст из чека — в БД хранится в `raw_text_debug` (Tesseract)."""
    if not op.ocr_data:
        return ""
    if isinstance(op.ocr_data, dict):
        return (
            op.ocr_data.get("raw_text_debug")
            or op.ocr_data.get("raw_text")
            or op.ocr_data.get("text")
            or ""
        )
    return str(op.ocr_data)


def _operation_row(db, op):
    api = op.api_data if isinstance(op.api_data, dict) else {}
    row_inner = api.get("row") if isinstance(api.get("row"), dict) else {}
    card_o = api.get("card") if isinstance(api.get("card"), dict) else {}
    ocr = op.ocr_data if isinstance(op.ocr_data, dict) else {}

    if op.source == "personal_receipt":
        card = "—"
        pname = ocr.get("fuel_type") or "—"
        pq = ocr.get("quantity") if ocr.get("quantity") not in (None, "") else "—"
        cost = ocr.get("total_sum") if ocr.get("total_sum") not in (None, "") else "—"
        azs = ocr.get("azs_number") or "—"
        car_api = op.actual_car or "—"
        drv = "—"
    else:
        card = api.get("cardNumber") or card_o.get("cardNumber") or "—"
        pname = api.get("productName") or row_inner.get("productName") or "—"
        pq = api.get("productQuantity") or row_inner.get("productQuantity") or 0
        cost = api.get("productCost") or row_inner.get("productCost") or 0
        azs = api.get("azsNumber") or row_inner.get("azsNumber") or row_inner.get("AzsCode") or "—"
        car_api = op.car_from_api or api.get("carNum") or row_inner.get("carNum") or "—"
        drv = api.get("driverName") or row_inner.get("driverName") or "—"

    presumed = db.query(User).filter_by(id=op.presumed_user_id).first() if op.presumed_user_id else None
    confirmed = db.query(User).filter_by(id=op.confirmed_user_id).first() if op.confirmed_user_id else None

    dt = op.date_time
    fuel_type = "Топливная карта" if op.source == "api" else "Личные средства"
    # Для листа "личные средства":
    # - "Предполагаемый" всегда прочерк
    # - "Фактический" = отправитель чека (presumed_user_id)
    # - "Кто окончательно подтвердил" = прочерк
    # - "Дата подтверждения" = прочерк
    if op.source == "personal_receipt":
        presumed_name = "—"
        factual_name = presumed.full_name if presumed else "—"
        final_confirmer = "—"
        confirmed_dt = "—"
    else:
        presumed_name = presumed.full_name if presumed else "—"
        factual_name = confirmed.full_name if confirmed else "—"
        final_confirmer = confirmed.full_name if confirmed else "—"
        confirmed_dt = op.confirmed_at.strftime("%d.%m.%Y %H:%M") if op.confirmed_at else "—"

    # Возвращаем список из 23 колонок согласно ТЗ
    return [
        op.id,                                      # ID
        fuel_type,                                  # Тип
        op.source or "—",                           # Источник
        dt.strftime("%d.%m.%Y") if dt else "—",    # Дата
        dt.strftime("%H:%M:%S") if dt else "—",    # Время
        card, pname, pq, cost, azs,                 # Карта, Топливо, Литры, Сумма, АЗС
        op.doc_number or "—",                       # Чек
        car_api, drv,                               # Авто(API), Водитель(API)
        _ocr_text(op),                              # Данные OCR
        presumed_name,                               # Предполагаемый
        factual_name,                                # Фактический
        op.actual_car or car_api,                   # Факт.Авто
        _first_confirmation_sender_name(db, op.id),
        final_confirmer,                             # Окончательно подтвердил
        STATUS_RU.get(op.status, op.status or "—"), # Статус (РУССКИЙ)
        confirmed_dt,                                # Дата подтв
        "",                                         # Примечание
        "Да" if op.status == "confirmed" else "Нет" # Готовность
    ]


def _ensure_workbook(path: Path) -> Workbook:
    """Гарантирует существование файла и ВСЕХ листов в нем."""
    wb = None
    if path.exists():
        try:
            wb = load_workbook(path)
        except Exception as e:
            logger.error(f"Не удалось открыть существующий файл: {e}")

    if wb is None:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Проверяем каждый лист из списка необходимых
    required_sheets = [SHEET_CARDS, SHEET_PERSONAL, SHEET_DISPUTED, SHEET_REF]
    changed = False
    for name in required_sheets:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(HEADERS)
            changed = True

    if changed or not path.exists():
        wb.save(path)
    return wb


def _sheet_has_id(ws, op_id: int) -> bool:
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0] == op_id:
            return True
    return False


def export_to_excel_final(op_id: int) -> None:
    """
    Выгружает операцию в соответствующий лист.
    """
    with get_db_session() as db:
        op = db.query(FuelOperation).filter_by(id=op_id).first()
        if not op:
            return

        st = op.status or ""
        disputed = st in ("requires_manual", "rejected_by_other", "import_error", "disputed")

        if disputed:
            if getattr(op, "exported_disputed_excel", False):
                return
            sheet_name = SHEET_DISPUTED
        elif st == "confirmed":
            if getattr(op, "exported_to_excel", False):
                return
            sheet_name = SHEET_CARDS if op.source == "api" else SHEET_PERSONAL
        else:
            logger.debug("[excel] skip export op_id=%s status=%s", op_id, st)
            return

        row = _operation_row(db, op)

        try:
            wb = _ensure_workbook(MASTER_FILE)
            ws = wb[sheet_name]
            if not _sheet_has_id(ws, op_id):
                ws.append(row)
            wb.save(MASTER_FILE)

            if disputed:
                if hasattr(op, "exported_disputed_excel"):
                    op.exported_disputed_excel = True
            else:
                if hasattr(op, "exported_to_excel"):
                    op.exported_to_excel = True
                if hasattr(op, "ready_for_waybill"):
                    op.ready_for_waybill = True

            db.commit()
            logger.info("[excel] op_id=%s sheet=%s", op_id, sheet_name)
        except PermissionError as e:
            logger.error("[excel] файл Excel занят: %s", e)
            raise
        except Exception as e:
            logger.exception("[excel] ошибка записи op_id=%s: %s", op_id, e)
            raise


def export_operation_to_excel(operation_id: int):
    with get_db_session() as db:
        op = db.query(FuelOperation).get(operation_id)
        if not op:
            return

        st = op.status

        # 1. Проверяем, является ли заправка спорной
        is_disputed = st in ("requires_manual", "rejected_by_other", "import_error", "disputed")

        if is_disputed:
            if getattr(op, "exported_disputed_excel", False):
                return
            sheet_name = SHEET_DISPUTED

        elif st == "confirmed":
            if getattr(op, "exported_to_excel", False):
                return
            sheet_name = SHEET_CARDS if op.source == "api" else SHEET_PERSONAL

        else:
            # Пропускаем статусы loaded, pending и т.д.
            return

        row_data = _operation_row(db, op)

        try:
            wb = _ensure_workbook(MASTER_FILE)
            ws = wb[sheet_name]

            # Добавляем строку
            ws.append(row_data)

            # Ставим отметку об экспорте
            if is_disputed:
                op.exported_disputed_excel = True
            else:
                op.exported_to_excel = True

            wb.save(MASTER_FILE)
            db.commit()
            logger.info(f"✅ Операция {operation_id} записана в лист {sheet_name}")

        except Exception as e:
            logger.error(f"❌ Ошибка экспорта в Excel: {e}")