"""
Полный отчёт Excel — та же структура листов и строк, что в bot/handlers/admin_import.btn_export_excel.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy.orm import Session

from src.app.models import FuelOperation, User


def build_full_fuel_report_excel(db: Session) -> tuple[BytesIO, int] | tuple[None, int]:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_cards = wb.create_sheet("Заправки_по_картам")
    ws_personal = wb.create_sheet("Заправки_личные_средства")
    ws_disputed = wb.create_sheet("Спорные заправки")

    headers = [
        "ID", "Тип", "Источник", "Дата", "Время", "Карта", "Топливо", "Литры", "Сумма", "АЗС", "Чек",
        "Авто(API)", "Водитель(API)", "OCR", "Предполагаемый", "Фактический", "Факт.Авто",
        "Инициатор", "Подтвердил", "Статус", "Дата подтв", "Прим", "Готовность",
    ]

    for ws in (ws_cards, ws_personal, ws_disputed):
        ws.append(headers)

    status_ru = {
        "loaded_from_api": "Загружена из API",
        "pending": "Ожидает подтверждения",
        "confirmed": "Подтверждена",
        "requires_manual": "Требует ручной обработки",
        "disputed": "Спорная",
        "rejected_by_other": "Отклонена",
        "import_error": "Ошибка импорта",
        "loaded": "Загружена",
        "awaiting_user_confirmation": "Ожидает подтверждения",
        "new": "Новая",
        "rejected": "Отклонена",
    }

    operations = db.query(FuelOperation).order_by(FuelOperation.date_time.desc()).all()

    if not operations:
        return None, 0

    for op in operations:
        api = op.api_data if isinstance(op.api_data, dict) else {}
        row_inner = api.get("row") if isinstance(api.get("row"), dict) else {}
        card_o = api.get("card") if isinstance(api.get("card"), dict) else {}

        card = api.get("cardNumber") or card_o.get("cardNumber") or "—"
        pname = api.get("productName") or row_inner.get("productName") or "—"
        pq = api.get("productQuantity") or row_inner.get("productQuantity") or 0
        cost = api.get("productCost") or row_inner.get("productCost") or 0
        azs = api.get("azsNumber") or row_inner.get("azsNumber") or row_inner.get("AzsCode") or "—"
        car_api = op.car_from_api or api.get("carNum") or row_inner.get("carNum") or "—"
        drv = api.get("driverName") or row_inner.get("driverName") or "—"

        presumed = db.query(User).filter_by(id=op.presumed_user_id).first() if op.presumed_user_id else None
        confirmed = db.query(User).filter_by(id=op.confirmed_user_id).first() if op.confirmed_user_id else None

        fuel_type = "Топливная карта" if op.source == "api" else "Личные средства"
        dt = op.date_time

        presumed_name = presumed.full_name if presumed else "—"
        confirmed_name = confirmed.full_name if confirmed else "—"
        current_status = status_ru.get(op.status, op.status or "—")

        row_data = [
            op.id,
            fuel_type,
            op.source or "api",
            dt.strftime("%d.%m.%Y") if dt else "—",
            dt.strftime("%H:%M:%S") if dt else "—",
            card,
            pname,
            pq,
            cost,
            azs,
            op.doc_number or "—",
            car_api,
            drv,
            "",
            presumed_name,
            confirmed_name,
            op.actual_car or car_api,
            presumed_name,
            confirmed_name,
            current_status,
            op.confirmed_at.strftime("%d.%m.%Y %H:%M") if op.confirmed_at else "—",
            "",
            "Да" if op.status == "confirmed" else "Нет",
        ]

        if op.status in ("requires_manual", "rejected_by_other", "disputed", "import_error"):
            ws_disputed.append(row_data)
        else:
            if op.source == "api":
                ws_cards.append(row_data)
            else:
                ws_personal.append(row_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(operations)
